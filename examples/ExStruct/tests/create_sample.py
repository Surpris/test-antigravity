from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

ws["A1"] = "Header 1"
ws["B1"] = "Header 2"
ws["A2"] = "Value 1"
ws["B2"] = "Value 2"

# Create a merged cell
ws.merge_cells("C1:D2")
ws["C1"] = "Merged Cell"

# Output to samples directory
output_dir = os.path.join(os.path.dirname(__file__), "../samples")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "sample.xlsx")
wb.save(output_path)
print(f"Created {output_path}")
